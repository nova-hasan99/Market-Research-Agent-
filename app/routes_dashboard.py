"""
Dashboard routes for MarketLens.
Handles user analysis history, admin panel, and export.
"""
import io
import json
import csv as csv_mod
from typing import Optional

from fastapi import APIRouter, Request, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse

from app.auth import (
    get_current_user,
    require_user, require_admin,
    require_user_api, require_admin_api,
)
from app.db import get_admin_client
from app.deps import templates

router = APIRouter()


# ─── helpers ──────────────────────────────────────────────────────────────────
def _admin_client():
    return get_admin_client()


def _fetch_user_analyses(user_id: str) -> list:
    client = _admin_client()
    result = (
        client.table("analyses")
        .select("id,asset_type,asset,score,bias,last_price,created_at")
        .eq("user_id", user_id)
        .order("created_at", desc=True)
        .execute()
    )
    return result.data or []


# ─── GET /dashboard ───────────────────────────────────────────────────────────
@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    from fastapi.responses import RedirectResponse
    result = await require_user(request)
    if isinstance(result, RedirectResponse):
        return result
    user = result

    # Support admin viewing another user's dashboard
    view_user_id = request.query_params.get("user_id", user["id"])
    if view_user_id != user["id"] and not user.get("is_admin"):
        view_user_id = user["id"]

    analyses = _fetch_user_analyses(view_user_id)

    forex_analyses = [a for a in analyses if a.get("asset_type") == "forex"]
    stock_analyses = [a for a in analyses if a.get("asset_type") == "stock"]

    return templates.TemplateResponse(request, "dashboard.html", {
        "user":            user,
        "forex_analyses":  forex_analyses,
        "stock_analyses":  stock_analyses,
        "view_user_id":    view_user_id,
    })


# ─── GET /admin/users — dedicated Super Admin user-management page ────────────
@router.get("/admin/users", response_class=HTMLResponse)
async def admin_users_page(request: Request):
    user = await get_current_user(request)
    if not user:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login?next=/admin/users", status_code=302)
    if not user.get("is_admin"):
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/dashboard", status_code=302)
    return templates.TemplateResponse(request, "admin_users.html", {"user": user})


# ─── GET /dashboard/view/{id} — full-page analysis viewer ────────────────────
@router.get("/dashboard/view/{analysis_id}", response_class=HTMLResponse)
async def view_analysis_page(request: Request, analysis_id: str):
    from fastapi.responses import RedirectResponse
    result = await require_user(request)
    if isinstance(result, RedirectResponse):
        return result
    user = result
    client = _admin_client()
    result = (
        client.table("analyses")
        .select("*")
        .eq("id", analysis_id)
        .maybe_single()
        .execute()
    )
    row = result.data
    if not row:
        raise HTTPException(404, "Analysis not found")
    if row["user_id"] != user["id"] and not user.get("is_admin"):
        raise HTTPException(403, "Forbidden")

    analysis_data = row.get("data") or {}
    return templates.TemplateResponse(request, "analysis_view.html", {
        "user":               user,
        "analysis":           row,
        "analysis_data_json": json.dumps(analysis_data, default=str),
    })


# ─── GET /dashboard/analysis/{id} — JSON API for JS ──────────────────────────
@router.get("/dashboard/analysis/{analysis_id}")
async def get_analysis(request: Request, analysis_id: str):
    user = await require_user_api(request)
    client = _admin_client()

    result = (
        client.table("analyses")
        .select("*")
        .eq("id", analysis_id)
        .maybe_single()
        .execute()
    )
    row = result.data
    if not row:
        raise HTTPException(404, "Analysis not found")
    if row["user_id"] != user["id"] and not user.get("is_admin"):
        raise HTTPException(403, "Forbidden")

    # Return the full data dict for JS renderResults()
    data = row.get("data") or {}
    return JSONResponse(data)


# ─── POST /api/dashboard/save ─────────────────────────────────────────────────
@router.post("/api/dashboard/save")
async def save_analysis(request: Request):
    user = await require_user_api(request)
    body = await request.json()

    client = _admin_client()
    result = (
        client.table("analyses")
        .insert({
            "user_id":    user["id"],
            "asset_type": body.get("asset_type", ""),
            "asset":      body.get("asset", ""),
            "score":      body.get("score"),
            "bias":       body.get("bias", ""),
            "last_price": str(body.get("last_price", "")),
            "data":       body,
        })
        .execute()
    )
    rows = result.data or []
    new_id = rows[0]["id"] if rows else None
    return {"ok": True, "id": new_id}


# ─── DELETE /api/dashboard/analyses/{id} ──────────────────────────────────────
@router.delete("/api/dashboard/analyses/{analysis_id}")
async def delete_analysis(request: Request, analysis_id: str):
    user = await require_user_api(request)
    client = _admin_client()

    row = (
        client.table("analyses")
        .select("user_id")
        .eq("id", analysis_id)
        .maybe_single()
        .execute()
    ).data
    if not row:
        raise HTTPException(404, "Not found")
    if row["user_id"] != user["id"] and not user.get("is_admin"):
        raise HTTPException(403, "Forbidden")

    client.table("analyses").delete().eq("id", analysis_id).execute()
    return {"ok": True}


# ─── DELETE /api/dashboard/analyses (bulk) ────────────────────────────────────
@router.delete("/api/dashboard/analyses")
async def delete_analyses_bulk(request: Request):
    user = await require_user_api(request)
    body = await request.json()
    ids  = body.get("ids", [])
    if not ids:
        return {"ok": True, "deleted": 0}

    client = _admin_client()
    # Only delete rows that belong to the user (or admin)
    q = client.table("analyses").delete().in_("id", ids)
    if not user.get("is_admin"):
        q = q.eq("user_id", user["id"])
    q.execute()
    return {"ok": True, "deleted": len(ids)}


# ─── GET /api/dashboard/export ────────────────────────────────────────────────
@router.get("/api/dashboard/export")
async def export_analyses(
    request:    Request,
    format:     str = "csv",
    ids:        Optional[str] = None,
    asset_type: Optional[str] = None,
):
    user = await require_user_api(request)
    client = _admin_client()

    q = client.table("analyses").select("*").eq("user_id", user["id"])
    if ids:
        id_list = [i.strip() for i in ids.split(",") if i.strip()]
        q = q.in_("id", id_list)
    if asset_type:
        q = q.eq("asset_type", asset_type)

    rows = (q.order("created_at", desc=True).execute()).data or []

    if format == "xlsx":
        return _export_xlsx(rows)
    # default: csv
    return _export_csv(rows)


def _flat_row(row: dict) -> dict:
    data       = row.get("data") or {}
    sentiment  = data.get("sentiment") or {}
    ai_sum     = data.get("ai_summary") or {}
    ks         = data.get("key_signal") or {}
    bias_raw   = row.get("bias", "") or data.get("bias", "")
    bias_label = "Bullish" if bias_raw == "up" else ("Bearish" if bias_raw == "down" else "Unclear")
    created    = row.get("created_at", "")
    date_str   = created[:16].replace("T", " ") if created else ""
    timeframe  = data.get("timeframe_label") or (data.get("timeframe") or "").replace("-", " ").title()
    key_signal_text = ks.get("text", "") if isinstance(ks, dict) else str(ks or "")
    vr = data.get("volatility_regime") or ""
    volatility = vr.get("label", "") if isinstance(vr, dict) else str(vr)
    try:
        sent_score = round(float(sentiment.get("score", 0)), 3)
    except (TypeError, ValueError):
        sent_score = ""

    return {
        "date":           date_str,
        "asset_type":     (row.get("asset_type") or "").title(),
        "asset":          row.get("asset", "") or data.get("asset", ""),
        "score":          row.get("score", "") if row.get("score") is not None else data.get("score", ""),
        "bias":           bias_label,
        "last_price":     row.get("last_price", "") or data.get("last_price", ""),
        "timeframe":      timeframe,
        "volatility":     volatility,
        "key_signal":     key_signal_text,
        "main_risk":      data.get("main_risk", ""),
        "sentiment":      (sentiment.get("label") or "").title(),
        "sent_score":     sent_score,
        "news_articles":  sentiment.get("article_count", ""),
        "ai_summary":     ai_sum.get("text", ""),
        "ai_provider":    ai_sum.get("provider", ""),
        "id":             row.get("id", ""),
    }


# Column spec: (display header, flat_row key, xlsx column width)
_EXPORT_COLS = [
    ("Date / Time",     "date",          20),
    ("Type",            "asset_type",     8),
    ("Asset",           "asset",         12),
    ("Score",           "score",          7),
    ("Bias",            "bias",          10),
    ("Last Price",      "last_price",    12),
    ("Timeframe",       "timeframe",     20),
    ("Volatility",      "volatility",    20),
    ("Key Signal",      "key_signal",    42),
    ("Main Risk",       "main_risk",     42),
    ("Sentiment",       "sentiment",     12),
    ("Sent. Score",     "sent_score",    12),
    ("News Articles",   "news_articles", 13),
    ("AI Summary",      "ai_summary",    62),
    ("AI Provider",     "ai_provider",   14),
    ("ID",              "id",            38),
]


def _export_csv(rows: list) -> StreamingResponse:
    from datetime import datetime as _dt, timezone as _tz
    output = io.StringIO()
    writer = csv_mod.writer(output)
    writer.writerow([h for h, _, _ in _EXPORT_COLS])
    for row in rows:
        flat = _flat_row(row)
        writer.writerow([flat.get(k, "") for _, k, _ in _EXPORT_COLS])
    output.seek(0)
    fname = f"marketlens_analyses_{_dt.now(_tz.utc).strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _export_xlsx(rows: list):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    from datetime import datetime as _dt, timezone as _tz

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analyses"

        num_cols = len(_EXPORT_COLS)
        thin     = Side(style="thin", color="1E2D3D")
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Row 1: title banner ──────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        tc = ws.cell(row=1, column=1)
        now_str  = _dt.now(_tz.utc).strftime("%Y-%m-%d %H:%M")
        tc.value = f"MarketLens - Analysis Export | {now_str} UTC | {len(rows)} record(s)"
        tc.font      = Font(bold=True, size=12, color="FF93C5FD", name="Calibri")
        tc.fill      = PatternFill("solid", fgColor="FF0D1321")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Row 2: column headers ─────────────────────────────────────────────
        hdr_fill = PatternFill("solid", fgColor="FF1E2D3D")
        hdr_font = Font(bold=True, color="FF93C5FD", name="Calibri", size=10)

        for ci, (label, _, width) in enumerate(_EXPORT_COLS, 1):
            cell = ws.cell(row=2, column=ci, value=label)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[2].height = 22

        # ── Data rows ─────────────────────────────────────────────────────────
        fill_odd  = PatternFill("solid", fgColor="FF0F1923")
        fill_even = PatternFill("solid", fgColor="FF111827")
        text_font = Font(name="Calibri", size=10, color="FFD1D5DB")
        wrap_font = Font(name="Calibri", size=9,  color="FFD1D5DB")

        score_styles = {
            "high": ("FF065F46", "FF34D399"),
            "mid":  ("FF78350F", "FFFCD34D"),
            "low":  ("FF7F1D1D", "FFF87171"),
        }
        bias_styles = {
            "Bullish": ("FF064E3B", "FF6EE7B7"),
            "Bearish": ("FF7F1D1D", "FFFCA5A5"),
            "Unclear": ("FF1F2937", "FF9CA3AF"),
        }

        for ri, row in enumerate(rows, 3):
            flat      = _flat_row(row)
            base_fill = fill_odd if ri % 2 == 1 else fill_even

            for ci, (_, key, _) in enumerate(_EXPORT_COLS, 1):
                val = flat.get(key, "")
                if isinstance(val, (dict, list)):
                    val = str(val)
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = border

                if key == "score":
                    try:
                        s = int(float(val))
                    except (TypeError, ValueError):
                        s = 0
                    style = "high" if s >= 70 else ("mid" if s >= 50 else "low")
                    bg, fg = score_styles[style]
                    cell.fill      = PatternFill("solid", fgColor=bg)
                    cell.font      = Font(bold=True, color=fg, name="Calibri", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="top")

                elif key == "bias":
                    bg, fg = bias_styles.get(str(val), ("FF1F2937", "FF9CA3AF"))
                    cell.fill      = PatternFill("solid", fgColor=bg)
                    cell.font      = Font(bold=True, color=fg, name="Calibri", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="top")

                elif key in ("ai_summary", "key_signal", "main_risk"):
                    cell.fill      = base_fill
                    cell.font      = wrap_font
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                else:
                    cell.fill      = base_fill
                    cell.font      = text_font
                    cell.alignment = Alignment(horizontal="left", vertical="top")

            summary_len = len(str(flat.get("ai_summary", "")))
            ws.row_dimensions[ri].height = max(15, min(int(summary_len / 55) * 14 + 15, 140))

        # ── Freeze header + auto-filter ────────────────────────────────────────
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"marketlens_analyses_{_dt.now(_tz.utc).strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )

    except Exception as exc:
        raise HTTPException(500, f"XLSX generation failed: {exc}") from exc


# ─── Admin: GET /api/admin/users ──────────────────────────────────────────────
@router.get("/api/admin/users")
async def admin_get_users(request: Request):
    await require_admin_api(request)
    client = _admin_client()

    profiles = (
        client.table("profiles")
        .select("id,name,email,is_admin,is_active,created_at,last_seen")
        .execute()
    ).data or []

    # Get analysis counts per user
    counts_raw = (
        client.table("analyses")
        .select("user_id")
        .execute()
    ).data or []

    count_map: dict = {}
    for row in counts_raw:
        uid = row["user_id"]
        count_map[uid] = count_map.get(uid, 0) + 1

    for p in profiles:
        p["analysis_count"] = count_map.get(p["id"], 0)

    return profiles


# ─── Email helper (reuse from routes_auth) ───────────────────────────────────
def _notify_user(email: str, name: str, subject: str, html: str) -> None:
    from app.routes_auth import _send_email
    _send_email(email, subject, html)


def _deactivate_email_html(name: str) -> str:
    from app.config import SITE_URL
    return f"""
    <div style="font-family:Inter,sans-serif;max-width:480px;margin:auto;
                background:#0d1117;color:#f0f6fc;border-radius:12px;padding:2rem">
      <h2 style="color:#f59e0b;margin-top:0">⚠️ Analysis Access Limited</h2>
      <p>Hi <strong>{name}</strong>,</p>
      <p>Due to high traffic and API rate limitations, new Forex and Stock analyses are <strong>temporarily unavailable</strong>.
         You can still <strong>login and view your dashboard</strong> with all your previous analyses.</p>
      <p style="margin-top:1rem;padding:1rem;background:rgba(245,158,11,0.15);border-left:3px solid #f59e0b">
         <strong>What you can do:</strong>
         <ul style="margin:0.5rem 0 0 1.5rem;padding:0">
           <li>View and export all your previous analyses</li>
           <li>Access your dashboard and profile</li>
           <li>Try running new analyses after a few hours when traffic decreases</li>
         </ul>
      </p>
      <p style="margin-top:1rem">If you need immediate assistance, please reach out to <strong>hasan@latticecode.pro</strong>.</p>
      <p style="color:#6b7280;font-size:0.85rem;margin-top:1.5rem">— The MarketLens Team</p>
    </div>"""


def _activate_email_html(name: str) -> str:
    from app.config import SITE_URL
    return f"""
    <div style="font-family:Inter,sans-serif;max-width:480px;margin:auto;
                background:#0d1117;color:#f0f6fc;border-radius:12px;padding:2rem">
      <h2 style="color:#10b981;margin-top:0">Account Reinstated</h2>
      <p>Hi <strong>{name}</strong>,</p>
      <p>Your MarketLens account has been <strong>reactivated</strong>.
         You can now log in and use all features as normal.</p>
      <a href="{SITE_URL}/login" style="display:inline-block;margin-top:1rem;
         padding:0.65rem 1.5rem;background:#10b981;color:#fff;border-radius:8px;
         text-decoration:none;font-weight:700">Sign In Now</a>
      <p style="color:#6b7280;font-size:0.85rem;margin-top:1.5rem">— The MarketLens Team</p>
    </div>"""


def _get_user_profile(uid: str) -> dict:
    """Fetch user profile (name + email) for email notifications."""
    try:
        client = _admin_client()
        rows = client.table("profiles").select("name,email").eq("id", uid).execute().data
        return rows[0] if rows else {}
    except Exception:
        return {}


# ─── Admin: POST /api/admin/users/{uid}/deactivate ───────────────────────────
@router.post("/api/admin/users/{uid}/deactivate")
async def admin_deactivate_user(request: Request, uid: str):
    await require_admin_api(request)
    client = _admin_client()
    client.table("profiles").update({"is_active": False}).eq("id", uid).execute()
    # Note: Not banning in Supabase Auth — allows user to still login but cannot analyze
    # Email notification (non-blocking)
    profile = _get_user_profile(uid)
    if profile.get("email"):
        import asyncio
        asyncio.get_event_loop().run_in_executor(
            None, _notify_user,
            profile["email"], profile.get("name", "User"),
            "Your MarketLens Account Has Been Suspended",
            _deactivate_email_html(profile.get("name", "User")),
        )
    return {"ok": True}


# ─── Admin: POST /api/admin/users/{uid}/activate ─────────────────────────────
@router.post("/api/admin/users/{uid}/activate")
async def admin_activate_user(request: Request, uid: str):
    await require_admin_api(request)
    client = _admin_client()
    client.table("profiles").update({"is_active": True}).eq("id", uid).execute()
    # Unban in Supabase Auth
    try:
        client.auth.admin.update_user_by_id(uid, {"ban_duration": "none"})
    except Exception:
        pass
    # Email notification
    profile = _get_user_profile(uid)
    if profile.get("email"):
        import asyncio
        asyncio.get_event_loop().run_in_executor(
            None, _notify_user,
            profile["email"], profile.get("name", "User"),
            "Your MarketLens Account Has Been Reinstated",
            _activate_email_html(profile.get("name", "User")),
        )
    return {"ok": True}


# ─── Admin: DELETE /api/admin/users/{uid} ────────────────────────────────────
@router.delete("/api/admin/users/{uid}")
async def admin_delete_user(request: Request, uid: str):
    """
    Permanently delete a user and ALL their data:
      1. Delete all analyses from the analyses table
      2. Delete profile row
      3. Delete from Supabase Auth (revokes all sessions)
    """
    me = await require_admin_api(request)
    if me["id"] == uid:
        raise HTTPException(400, "Cannot delete your own account")

    client = _admin_client()
    # 1. Delete all analyses
    try:
        client.table("analyses").delete().eq("user_id", uid).execute()
    except Exception:
        pass
    # 2. Delete profile row
    try:
        client.table("profiles").delete().eq("id", uid).execute()
    except Exception:
        pass
    # 3. Delete from Supabase Auth
    try:
        client.auth.admin.delete_user(uid)
    except Exception as e:
        raise HTTPException(500, f"Auth deletion failed: {e}")

    return {"ok": True, "deleted": uid}


# ─── Admin: GET /api/admin/users/{uid}/analyses ──────────────────────────────
@router.get("/api/admin/users/{uid}/analyses")
async def admin_get_user_analyses(request: Request, uid: str):
    await require_admin_api(request)
    rows = _fetch_user_analyses(uid)
    return rows
